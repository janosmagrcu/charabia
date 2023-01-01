import os
import pickle
from openpyxl import load_workbook
from num2words import num2words

"""
Création d'un fichier pickle contenant un dictionnaire réduit à partir du Lexicon
col 1 - orthographe
col 2 - gram_genre_nombre_infover

Prend une trentaine de secondes
"""

local_path = os.path.dirname(__file__)
wb = load_workbook(local_path + "/Data/Lexique383 - propre.xlsx")
sheet = wb.active

lexicon = {}

# liste des mots à retirer du lexicon :
retirer = {
	"sécolle",
	"lez",
	"s'",
	"l'",
	"d'",
	"t'",
	"n'",
	"la-la-la",
	"lès",
	"to",
	"ès",
	"1",
	"0",
	'ains',
  'please',
  'com',
  'lez',
  'milli',
  'seq',
  'fors',
  'mn',
  'lambi',
  'tiens-la-moi',
  'port-salut',
  'all right',
  'abc',
  'jà'
	}


for i,row in enumerate(sheet.iter_rows(min_row=2)):
	ortho, gram, genre, nombre, infover = row
	ort,gra,gen,nbr,ivs = ortho.value,gram.value,genre.value,nombre.value,infover.value

	if ort in retirer: continue
	
	ivs = {iv for iv in ivs.split(';')if iv} if ivs else ('',)

	for iv in ivs:
		if iv=='' or 'par' in iv:
			g = '_'.join([e for e in (gra,gen,nbr,iv) if e])
		else:
			g = '_'.join([e for e in (gra,iv) if e])

		lexicon[ort] = lexicon.get(ort,[]) + [g]


# ajout des NUM et ORD de 0 à 1000:

for n in range(1001):
	lexicon[num2words(n, lang='fr')] = ['NUM']
	lexicon[num2words(n, lang='fr', to='ordinal')] = ['ORD']


# apport de corrections manuelles:
corrections = {
	"connais": ['VER_imp:pre:2s', 'VER_ind:pre:2s', 'VER_ind:pre:1s'],
	"ira": ['VER_ind:fut:3s'],

	"leurs": ['ADJ:pos_p'],
	"leur": ['ADJ:pos_s'],
	"notre": ['ADJ:pos_s'],
	"mon": ['ADJ:pos_m_s'],
	"ma": ['ADJ:pos_f_s'],
	"mes": ['ADJ:pos_p'],
	"ses": ['ADJ:pos_p'],
	"votre": ['ADJ:pos_s'],
	"sa": ['ADJ:pos_f_s'],
	"ses": ['ADJ:pos_p'],
	"son": ['ADJ:pos_m_s','NOM_m_s'],
	"ta": ['ADJ:pos_f_s'],
	"tes": ['ADJ:pos_p'],
	"ton": ['ADJ:pos_m_s','NOM_m_s'],
	"vos": ['ADJ:pos_p'],
	"votre": ['ADJ:pos_s'],



	"mien": ['PRO:pos_m_s'],
	"mienne": ['PRO:pos_f_s'],
	"miennes": ['PRO:pos_f_p'],
	"miens": ['PRO:pos_m_p'],
	"nôtre": ['PRO:pos_s'],
	"nôtres": ['PRO:pos_p'],
	"sien": ['PRO:pos_m_s'],
	"sienne": ['PRO:pos_f_s'],
	"siennes": ['PRO:pos_f_p'],
	"siens": ['PRO:pos_m_p'],
	"tien": ['PRO:pos_m_s'],
	"tienne": ['PRO:pos_f_s'],
	"tiennes": ['PRO:pos_f_p'],
	"tiens": ['PRO:pos_m_p'],
	"vôtre": ['PRO:pos_s'],
	"vôtres": ['PRO:pos_p'],

	"con": ['NOM_m_s','ADJ_m_s'],

}
lexicon.update(corrections)


# lexicon inverse :

lexicon_inv = {}

for mot,gram in lexicon.items():
	for g in gram:
		lexicon_inv[g] = lexicon_inv.get(g,set()) | {mot}

# pickle dump
with open(local_path + '/Data/Lexicon_simplifié.pkl', 'wb') as f:
    pickle.dump(lexicon, f)

with open(local_path + '/Data/Lexicon_inverse.pkl', 'wb') as f:
    pickle.dump(lexicon_inv, f)