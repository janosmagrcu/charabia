import os
import pickle
from openpyxl import load_workbook
from num2words import num2words

"""
Création d'un fichier pickle contenant un dictionnaire réduit à partir du Lexicon
col 1 - orthographe
col 2 - gram_genre_nombre_infover

Prend une trentaine de secondes
"""

local_path = os.path.dirname(__file__)
wb = load_workbook(local_path + "/Data/Lexique383 - avec fréquences.xlsx")
sheet = wb.active

lexicon = {}

# liste des mots à retirer du lexicon :
retirer = {
	"sécolle",
	"lez",
	"s'",
	"l'",
	"d'",
	"t'",
	"n'",
	"la-la-la",
	"lès",
	"to",
	"ès",
	"1",
	"0",
	'ains',
	'please',
	'com',
	'lez',
	'milli',
	'seq',
	'fors',
	'mn',
	'lambi',
	'tiens-la-moi',
	'port-salut',
	'all right',
	'abc',
	'jà',
	'qua',
	'più',
	'pa'
	}


for i,row in enumerate(sheet.iter_rows(min_row=2)):
	ortho, gram, genre, nombre, infover, freq = row
	ort,gra,gen,nbr,ivs,frq = [e.value for e in (ortho, gram, genre, nombre, infover, freq)]

	if ort in retirer: continue
	
	ivs = {iv for iv in ivs.split(';')if iv} if ivs else ('',)

	for iv in ivs:
		if iv=='' or 'par' in iv:
			g = '_'.join([e for e in (gra,gen,nbr,iv) if e])
		else:
			g = '_'.join([e for e in (gra,iv) if e])

		lexicon[ort] = lexicon.get(ort,{}) | {g:frq}


# ajout des NUM et ORD de 0 à 1000:

n_max = 1000
for n in range(2,n_max+1): # pour ne pas toucher à "un"

	lexicon[num2words(n, lang='fr')] = {'ADJ:num':(n_max-n)**.25}
	lexicon[num2words(n, lang='fr', to='ordinal')] = {'ADJ':(n_max-n)**.4}


# apport de corrections manuelles:
corrections = {
	"connais": {'VER_imp:pre:2s':111, 'VER_ind:pre:2s':111, 'VER_ind:pre:1s':111},
	"ira": {'VER_ind:fut:3s':28},
	"o": {'NOM_m':11},
	"team": {'NOM_m_s':0.2},
	"à brûle-pourpoint": {'ADV':1.22},
	"à cloche-pied": {'ADV':1.35},
	"à rebrousse-poil": {'ADV':0.81},
	"à tire-larigot": {'ADV':0.14},
	"à touche-touche": {'ADV':0.14},
	"à tue-tête": {'ADV':3.72},

	"leurs": {'ADJ:pos_p':886},
	"leur": {'ADJ:pos_s':415},
	"notre": {'ADJ:pos_s':680},
	"mon": {'ADJ:pos_m_s':2307},
	"ma": {'ADJ:pos_f_s':1667},
	"mes": {'ADJ:pos_p':1023},
	"ses": {'ADJ:pos_p':3105},
	"votre": {'ADJ:pos_s':427},
	"sa": {'ADJ:pos_f_s':3732},
	"son": {'ADJ:pos_m_s':4696,'NOM_m_s':49},
	"ta": {'ADJ:pos_f_s':251},
	"tes": {'ADJ:pos_p':145},
	"ton": {'ADJ:pos_m_s':310,'NOM_m_s':138},
	"vos": {'ADJ:pos_p':180},
	"votre": {'ADJ:pos_s':427},

	# A vérifier, peuvent être pronom ou adj selon contexte
	# "mien": {'PRO:pos_m_s':36},
	# "mienne": {'PRO:pos_f_s':41},
	# "miennes": {'PRO:pos_f_p':9.6},
	# "miens": {'PRO:pos_m_p':21},
	# "nôtre": {'PRO:pos_s':680},
	# "nôtres": {'PRO:pos_p':300},
	# "sien": {'PRO:pos_m_s':},
	# "sienne": {'PRO:pos_f_s':},
	# "siennes": {'PRO:pos_f_p':},
	# "siens": {'PRO:pos_m_p':},
	# "tien": {'PRO:pos_m_s':},
	# "tienne": {'PRO:pos_f_s':},
	# "tiennes": {'PRO:pos_f_p':},
	# "tiens": {'PRO:pos_m_p':},
	# "vôtre": {'PRO:pos_s':},
	# "vôtres": {'PRO:pos_p':},

	"con": {'NOM_m_s':45,'ADJ_m_s':26},
}
lexicon.update(corrections)

detailed_gram_list = {g for grams in lexicon.values() for g in grams}
gram_list = {g.split('_')[0] for g in detailed_gram_list}
print(sorted(detailed_gram_list))
print(sorted(gram_list))
"""
['', 
'ADJ', 'ADJ:dem', 'ADJ:ind', 'ADJ:int', 'ADJ:num', 'ADJ:pos', 
'ADV', 
'ART:def', 'ART:ind', 
'AUX', 
'CON', 
'NOM', 
'NUM', 
'ONO', 
'ORD', 
'PRE', 
'PRO:dem', 'PRO:ind', 'PRO:int', 'PRO:per', 'PRO:pos', 'PRO:rel', 
'VER'] 
"""

# lexicon inverse :

lexicon_inv = {}

for mot,gram in lexicon.items():
	for g,frq in gram.items():
		lexicon_inv[g] = lexicon_inv.get(g,{}) | {mot:frq}

# pickle dump
with open(local_path + '/Data/Lexicon_simplifié.pkl', 'wb') as f:
    pickle.dump(lexicon, f)

with open(local_path + '/Data/Lexicon_inverse.pkl', 'wb') as f:
    pickle.dump(lexicon_inv, f)